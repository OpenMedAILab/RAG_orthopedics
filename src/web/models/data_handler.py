"""数据处理模块

提供数据管理、评分管理和 Excel 导出功能。
采用设计模式：观察者模式、策略模式、单例模式。

仅支持对比评分模式 (comparison) - 一次评分两个项目 (NoRAG vs RAG)
"""

import json
import logging
import math
from abc import ABC, abstractmethod
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

from config import Config
from models.heartbeat_manager import get_heartbeat_manager

# 配置日志
logger = logging.getLogger(__name__)


# ==================== 观察者模式 ====================

class Observer(ABC):
    """观察者抽象基类
    
    定义观察者的接口，当主题状态变化时接收通知。
    """

    @abstractmethod
    def update(self, subject: 'Subject', event: str, data: Optional[Any] = None) -> None:
        """当主题状态发生变化时调用此方法
        
        Args:
            subject: 发生变化的主题对象
            event: 事件类型
            data: 事件相关数据
        """
        pass


class Subject(ABC):
    """主题抽象基类
    
    维护观察者列表，并提供添加、删除和通知观察者的方法。
    """

    def __init__(self) -> None:
        """初始化主题，创建空的观察者列表"""
        self._observers: List[Observer] = []

    def attach(self, observer: Observer) -> None:
        """添加观察者
        
        Args:
            observer: 要添加的观察者对象
        """
        self._observers.append(observer)

    def detach(self, observer: Observer) -> None:
        """移除观察者
        
        Args:
            observer: 要移除的观察者对象
        """
        self._observers.remove(observer)

    def notify(self, event: str, data: Optional[Any] = None) -> None:
        """通知所有观察者
        
        Args:
            event: 事件类型
            data: 事件相关数据
        """
        for observer in self._observers:
            observer.update(self, event, data)


# ==================== 策略模式 - JSON 解析 ====================

class JsonParserStrategy(ABC):
    """JSON 解析策略抽象基类
    
    定义 JSON 解析器的接口。
    """

    @abstractmethod
    def parse(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """解析 JSON 数据
        
        Args:
            data: 原始 JSON 数据
            
        Returns:
            Dict[str, Any]: 解析后的数据
        """
        pass


class ComparisonJsonParser(JsonParserStrategy):
    """对比评分模式 JSON 解析策略
    
    将原始数据转换为对比评分模式的标准格式。
    """

    def parse(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """解析对比评分 JSON 格式
        
        Args:
            data: 原始 JSON 数据
            
        Returns:
            Dict[str, Any]: 标准化后的对比评分数据
        """
        parsed_data = data.copy()
        
        # 标准化字段名
        if 'pairId' not in parsed_data:
            parsed_data['pairId'] = parsed_data.get('case_id', '')
        
        # 构建 caseInfo
        if 'caseInfo' not in parsed_data:
            parsed_data['caseInfo'] = {
                'caseId': parsed_data.get('case_id', ''),
                'caseNumber': parsed_data.get('case_number', ''),
                'context': parsed_data.get('context', ''),
                'physicalExamination': parsed_data.get('physical_examination', ''),
                'imagingExamination': parsed_data.get('imaging_examination', '')
            }
        
        # 标准化 model 字段
        if 'model' not in parsed_data:
            parsed_data['model'] = parsed_data.get('evaluated_model', '')
        
        # 构建 sideA (NoRAG)
        if 'sideA' not in parsed_data:
            parsed_data['sideA'] = {
                'sideId': Config.SIDE_A_ID,
                'mode': Config.SIDE_A_MODE,
                'label': Config.SIDE_A_LABEL,
                'qaPairs': parsed_data.get('qa_pairs', []),
                'ragContext': parsed_data.get('rag_context', ''),
                'evaluation': {}
            }
        
        # 构建 sideB (RAG)
        if 'sideB' not in parsed_data:
            parsed_data['sideB'] = {
                'sideId': Config.SIDE_B_ID,
                'mode': Config.SIDE_B_MODE,
                'label': Config.SIDE_B_LABEL,
                'qaPairs': [],
                'ragContext': '',
                'evaluation': {}
            }
        
        # 添加元数据
        if 'metadata' not in parsed_data:
            parsed_data['metadata'] = {}
        
        return parsed_data


class JsonParserFactory:
    """JSON 解析器工厂
    
    根据类型创建相应的解析器实例。
    """

    @staticmethod
    def create_parser(parser_type: str = 'comparison') -> JsonParserStrategy:
        """根据类型创建解析器实例
        
        Args:
            parser_type: 解析器类型，目前仅支持 'comparison'
            
        Returns:
            JsonParserStrategy: 解析器实例
        """
        if parser_type == 'comparison':
            return ComparisonJsonParser()
        return ComparisonJsonParser()


# ==================== 数据验证器 ====================

class DataValidator:
    """数据验证器
    
    验证数据格式是否正确。
    """

    @staticmethod
    def validate_pair_data(pair_data: Dict[str, Any]) -> bool:
        """验证对比评分对数据
        
        Args:
            pair_data: 对比评分对数据
            
        Returns:
            bool: 数据是否有效
        """
        required_fields = ['pairId', 'caseInfo', 'model', 'sideA', 'sideB']
        if not all(field in pair_data for field in required_fields):
            return False
        
        side_a = pair_data.get('sideA', {})
        side_b = pair_data.get('sideB', {})
        
        if 'sideId' not in side_a or 'qaPairs' not in side_a:
            return False
        if 'sideId' not in side_b or 'qaPairs' not in side_b:
            return False
        
        return True


class RatingValidator:
    """评分验证器
    
    验证评分数据是否有效。
    """

    @staticmethod
    def validate_scores(scores: Dict[str, Any]) -> Tuple[bool, str]:
        """验证评分值是否有效
        
        Args:
            scores: 评分字典，包含三个维度的评分
            
        Returns:
            Tuple[bool, str]: (是否有效，错误信息)
        """
        required_fields = Config.RATING_DIMENSIONS
        
        for field in required_fields:
            if field not in scores:
                return False, f"缺少评分维度：{field}"
            
            value = scores[field]
            if value is None:
                return False, f"评分维度 {field} 不能为空"
            if not isinstance(value, int) or not Config.validate_rating_score(value):
                return False, f"评分维度 {field} 的值必须在 {Config.RATING_MIN_SCORE}-{Config.RATING_MAX_SCORE} 之间"
        
        return True, ""

    @staticmethod
    def validate_rating_data(data: Dict[str, Any]) -> Tuple[bool, str]:
        """验证完整的评分数据
        
        Args:
            data: 评分数据字典
            
        Returns:
            Tuple[bool, str]: (是否有效，错误信息)
        """
        required_fields = ['index', 'user_id', 'side_a_scores', 'side_b_scores']
        
        for field in required_fields:
            if field not in data:
                return False, f"缺少必填字段：{field}"
        
        is_valid, error_msg = RatingValidator.validate_scores(data['side_a_scores'])
        if not is_valid:
            return False, f"Side A 评分无效：{error_msg}"
        
        is_valid, error_msg = RatingValidator.validate_scores(data['side_b_scores'])
        if not is_valid:
            return False, f"Side B 评分无效：{error_msg}"
        
        return True, ""


# ==================== 工具类 ====================

class NanConverter:
    """NaN 转换器
    
    将 NaN 值转换为 None，以便正确序列化为 JSON。
    """

    @staticmethod
    def convert(obj: Any) -> Any:
        """递归转换 NaN 值为 None
        
        Args:
            obj: 要转换的对象
            
        Returns:
            Any: 转换后的对象
        """
        if isinstance(obj, float) and math.isnan(obj):
            return None
        elif isinstance(obj, dict):
            return {key: NanConverter.convert(value) for key, value in obj.items()}
        elif isinstance(obj, list):
            return [NanConverter.convert(item) for item in obj]
        return obj


# ==================== 单例模式 - JSON 数据管理器 ====================

class JsonDataManager(Subject):
    """JSON 数据管理器 - 单例模式
    
    管理 JSON 数据的加载、保存和更新，支持观察者模式。
    """

    _instance: Optional['JsonDataManager'] = None

    def __new__(cls) -> 'JsonDataManager':
        """创建单例实例
        
        Returns:
            JsonDataManager: 单例实例
        """
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance._initialized = False
        return cls._instance

    def __init__(self) -> None:
        """初始化数据管理器"""
        if not self._initialized:
            super().__init__()
            self._json_data: List[Dict[str, Any]] = []
            self._parser_strategy: JsonParserStrategy = JsonParserFactory.create_parser('comparison')
            self._validator = DataValidator()
            self._initialized = True

    @property
    def mode(self) -> str:
        """获取当前模式
        
        Returns:
            str: 当前模式名称
        """
        return 'comparison'

    def set_parser_strategy(self, strategy: JsonParserStrategy) -> None:
        """设置解析策略
        
        Args:
            strategy: 新的解析策略
        """
        self._parser_strategy = strategy

    def load_json_data(self, file_path: Path, app: Optional[Any] = None) -> None:
        """加载 JSON 数据

        Args:
            file_path: JSON 文件路径
            app: Flask 应用实例，用于日志记录
        """
        try:
            if file_path.exists():
                with open(file_path, 'r', encoding='utf-8') as f:
                    raw_data = json.load(f)

                # 根据对比评分数据格式设计.md，JSON 文件格式为 {"ratingPairs": [...]}
                # 需要提取 ratingPairs 数组
                if isinstance(raw_data, dict):
                    # 支持多种可能的键名
                    if 'ratingPairs' in raw_data:
                        rating_pairs = raw_data['ratingPairs']
                    elif 'data' in raw_data:
                        rating_pairs = raw_data['data']
                    elif 'items' in raw_data:
                        rating_pairs = raw_data['items']
                    else:
                        # 如果都不是，尝试直接使用字典值
                        rating_pairs = list(raw_data.values())[0] if raw_data else []
                elif isinstance(raw_data, list):
                    # 兼容旧的数组格式
                    rating_pairs = raw_data
                else:
                    rating_pairs = []

                # 解析每个评分对
                self._json_data = [self._parser_strategy.parse(item) for item in rating_pairs]
                
                # 验证数据格式
                self._json_data = [
                    item for item in self._json_data
                    if DataValidator.validate_pair_data(item)
                ]

                if app:
                    app.logger.info(f"成功加载 JSON 数据，共 {len(self._json_data)} 条记录")
                else:
                    logger.info(f"成功加载 JSON 数据，共 {len(self._json_data)} 条记录")

                self.notify('json_loaded', {'count': len(self._json_data)})
            else:
                if app:
                    app.logger.error(f"错误：找不到 JSON 数据文件：{file_path}")
                else:
                    logger.error(f"错误：找不到 JSON 数据文件：{file_path}")
                self._json_data = []
        except json.JSONDecodeError as e:
            if app:
                app.logger.error(f"JSON 解析错误：{e}")
            else:
                logger.error(f"JSON 解析错误：{e}")
            self._json_data = []
        except Exception as e:
            if app:
                app.logger.error(f"加载 JSON 数据时发生错误：{e}")
            else:
                logger.error(f"加载 JSON 数据时发生错误：{e}")
            self._json_data = []

    def save_json_data(self, file_path: Path) -> None:
        """保存 JSON 数据

        Args:
            file_path: JSON 文件路径
        """
        try:
            # 根据对比评分数据格式设计.md，保存为 {"ratingPairs": [...]} 格式
            output_data = {"ratingPairs": self._json_data}
            
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(output_data, f, ensure_ascii=False, indent=2)
            self.notify('json_saved', {'file_path': str(file_path)})
            logger.info(f"成功保存 JSON 数据到：{file_path}")
        except Exception as e:
            logger.error(f"保存 JSON 数据失败：{e}")
            raise

    def get_json_data(self) -> List[Dict[str, Any]]:
        """获取 JSON 数据
        
        Returns:
            List[Dict[str, Any]]: JSON 数据列表
        """
        return self._json_data

    def set_json_data(self, data: List[Dict[str, Any]]) -> None:
        """设置 JSON 数据
        
        Args:
            data: 新的 JSON 数据列表
        """
        self._json_data = data
        self.notify('json_updated', {'count': len(data)})


# ==================== 评分管理器 ====================

class RatingManager:
    """评分管理器
    
    管理评分状态、评分保存和用户统计。
    """

    def __init__(self, user_manager: Any, data_manager: JsonDataManager) -> None:
        """初始化评分管理器
        
        Args:
            user_manager: 用户管理器实例
            data_manager: JSON 数据管理器实例
        """
        self.rating_status: Dict[int, Dict[str, Any]] = {}
        self.last_modified: Dict[int, Dict[str, Any]] = {}
        self.rating_logs: List[Dict[str, Any]] = []
        self.user_manager = user_manager
        self.data_manager = data_manager
        self.validator = RatingValidator()

    def get_all_ratings(self) -> List[Dict[str, Any]]:
        """获取所有评分项列表
        
        Returns:
            List[Dict[str, Any]]: 评分项列表
        """
        all_items = []
        json_data = self.data_manager.get_json_data()
        
        # 从心跳管理器获取所有状态
        heartbeat_manager = get_heartbeat_manager()

        for idx, item_data in enumerate(json_data):
            # 从心跳管理器获取状态（优先）
            heartbeat_status = heartbeat_manager.get_status(idx, '')
            
            # 如果没有心跳状态，使用 rating_status
            if heartbeat_status.get('status') == Config.RATING_STATUS_IDLE:
                status_info = self.rating_status.get(idx, {
                    'status': Config.RATING_STATUS_IDLE,
                    'user_id': None,
                    'timestamp': None
                })
                status = status_info['status']
                status_user = status_info.get('user_id')
            else:
                status = heartbeat_status.get('status', Config.RATING_STATUS_IDLE)
                status_user = heartbeat_status.get('locked_by', {}).get('user_id')

            # 检查两侧的评分状态
            side_a = item_data.get('sideA', {})
            side_b = item_data.get('sideB', {})
            eval_a = side_a.get('evaluation', {})
            eval_b = side_b.get('evaluation', {})
            scores_a = eval_a.get('scores', {})
            scores_b = eval_b.get('scores', {})

            side_a_rated = all(
                k in scores_a for k in Config.RATING_DIMENSIONS
            )
            side_b_rated = all(
                k in scores_b for k in Config.RATING_DIMENSIONS
            )

            item = {
                'index': idx + 1,
                'status': status,
                'status_user': status_user,
                'rated': side_a_rated and side_b_rated,
                'side_a_rated': side_a_rated,
                'side_b_rated': side_b_rated,
                'pair_id': item_data.get('pairId', ''),
                'case_info': item_data.get('caseInfo', {}),
                'model': item_data.get('model', ''),
            }
            all_items.append(NanConverter.convert(item))

        return all_items

    def save_rating(self, data: Dict[str, Any], app: Optional[Any] = None) -> Dict[str, Any]:
        """保存评分
        
        Args:
            data: 评分数据字典
            app: Flask 应用实例，用于日志记录
            
        Returns:
            Dict[str, Any]: 保存结果 {'success': bool, 'error': str}
        """
        index = data.get('index')
        user_id = data.get('user_id')
        rated_by_name = data.get('rated_by_name')

        # 验证评分数据
        is_valid, error_msg = self.validator.validate_rating_data(data)
        if not is_valid:
            return {"success": False, "error": error_msg}

        array_index = index - 1

        # 检查锁状态
        if array_index in self.rating_status:
            status = self.rating_status[array_index]
            if status.get('status') == Config.RATING_STATUS_PROCESSING:
                locked_by = status.get('user_id')
                if locked_by != user_id:
                    return {
                        "success": False, 
                        "error": f"评分项正在被用户 {locked_by} 处理"
                    }

        json_data = self.data_manager.get_json_data()
        if array_index >= len(json_data):
            return {"success": False, "error": "评分项索引超出范围"}

        item_data = json_data[array_index]

        # 提取评分数据
        side_a_scores = data.get('side_a_scores', {})
        side_a_reasoning = data.get('side_a_comment', '')
        side_b_scores = data.get('side_b_scores', {})
        side_b_reasoning = data.get('side_b_comment', '')

        # 更新 sideA
        self._update_side_evaluation(
            item_data, 'sideA', side_a_scores, side_a_reasoning, user_id, rated_by_name
        )
        
        # 更新 sideB
        self._update_side_evaluation(
            item_data, 'sideB', side_b_scores, side_b_reasoning, user_id, rated_by_name
        )

        # 更新状态
        self._update_rating_status(array_index, user_id)
        self._update_last_modified(array_index, user_id)
        self._add_rating_log(index, user_id, rated_by_name, data)

        # 准备 Excel 数据
        excel_data = self._prepare_excel_data(item_data, side_a_scores, side_b_scores, 
                                               side_a_reasoning, side_b_reasoning, rated_by_name)

        # 保存数据
        try:
            self.data_manager.set_json_data(json_data)
            self.data_manager.save_json_data(Config.JSON_DATA_FILE)
            
            if app:
                app.logger.info(f"成功保存评分：index={index}, user={user_id}")
            else:
                logger.info(f"成功保存评分：index={index}, user={user_id}")

            # 通知 Excel 观察者保存
            self.data_manager.notify('rating_saved', excel_data)
            
            return {"success": True}
        except Exception as e:
            logger.error(f"保存评分失败：{e}")
            return {"success": False, "error": f"保存失败：{str(e)}"}

    def _update_side_evaluation(
        self, 
        item_data: Dict[str, Any], 
        side_key: str, 
        scores: Dict[str, Any], 
        reasoning: str, 
        user_id: str, 
        rated_by_name: str
    ) -> None:
        """更新一侧的评分数据
        
        Args:
            item_data: 评分项数据
            side_key: 侧边键名 ('sideA' 或 'sideB')
            scores: 评分字典
            reasoning: 评分理由
            user_id: 用户 ID
            rated_by_name: 评分人姓名
        """
        if side_key not in item_data:
            item_data[side_key] = {}
        if 'evaluation' not in item_data[side_key]:
            item_data[side_key]['evaluation'] = {}

        item_data[side_key]['evaluation'] = {
            'scores': {
                'medical_accuracy': scores.get('medical_accuracy'),
                'key_point_recall': scores.get('key_point_recall'),
                'logical_completeness': scores.get('logical_completeness')
            },
            'reasoning': reasoning,
            'rated_by': user_id,
            'rated_by_name': rated_by_name,
            'rated_at': datetime.now().isoformat()
        }

    def _update_rating_status(self, array_index: int, user_id: str) -> None:
        """更新评分状态
        
        Args:
            array_index: 数组索引
            user_id: 用户 ID
        """
        self.rating_status[array_index] = {
            'status': Config.RATING_STATUS_COMPLETED,
            'user_id': user_id,
            'timestamp': pd.Timestamp.now()
        }

    def _update_last_modified(self, array_index: int, user_id: str) -> None:
        """更新最后修改信息
        
        Args:
            array_index: 数组索引
            user_id: 用户 ID
        """
        self.last_modified[array_index] = {
            'timestamp': pd.Timestamp.now(),
            'user_id': user_id
        }

    def _add_rating_log(
        self, 
        index: int, 
        user_id: str, 
        rated_by_name: str, 
        data: Dict[str, Any]
    ) -> None:
        """添加评分日志
        
        Args:
            index: 评分项索引
            user_id: 用户 ID
            rated_by_name: 评分人姓名
            data: 评分数据
        """
        self.rating_logs.append({
            'index': index,
            'user_id': user_id,
            'rated_by_name': rated_by_name,
            'timestamp': pd.Timestamp.now(),
            'changes': data
        })

    def _prepare_excel_data(
        self, 
        item_data: Dict[str, Any], 
        side_a_scores: Dict[str, Any], 
        side_b_scores: Dict[str, Any],
        side_a_reasoning: str,
        side_b_reasoning: str,
        rated_by_name: str
    ) -> Dict[str, Any]:
        """准备 Excel 数据
        
        Args:
            item_data: 评分项数据
            side_a_scores: Side A 评分
            side_b_scores: Side B 评分
            side_a_reasoning: Side A 评论
            side_b_reasoning: Side B 评论
            rated_by_name: 评分人姓名
            
        Returns:
            Dict[str, Any]: Excel 数据字典
        """
        return {
            'pair_id': item_data.get('pairId', ''),
            'case_number': item_data.get('caseInfo', {}).get('caseNumber', ''),
            'model': item_data.get('model', ''),
            'side_a_accuracy': side_a_scores.get('medical_accuracy'),
            'side_a_recall': side_a_scores.get('key_point_recall'),
            'side_a_logic': side_a_scores.get('logical_completeness'),
            'side_a_comment': side_a_reasoning,
            'side_b_accuracy': side_b_scores.get('medical_accuracy'),
            'side_b_recall': side_b_scores.get('key_point_recall'),
            'side_b_logic': side_b_scores.get('logical_completeness'),
            'side_b_comment': side_b_reasoning,
            'rated_by_name': rated_by_name
        }

    def get_rating_last_modified(self, index: int) -> Optional[Dict[str, Any]]:
        """获取评分项的最后修改信息
        
        Args:
            index: 评分项索引
            
        Returns:
            Optional[Dict[str, Any]]: 最后修改信息，不存在则返回 None
        """
        array_index = index - 1
        return self.last_modified.get(array_index)

    def get_rating_history(self, index: int) -> List[Dict[str, Any]]:
        """获取评分项的修改历史
        
        Args:
            index: 评分项索引
            
        Returns:
            List[Dict[str, Any]]: 修改历史列表
        """
        array_index = index - 1
        return [log for log in self.rating_logs if log['index'] - 1 == array_index]

    def set_rating_processing(self, index: int, user_id: str) -> Dict[str, Any]:
        """设置评分项为处理中状态
        
        Args:
            index: 评分项索引
            user_id: 用户 ID
            
        Returns:
            Dict[str, Any]: 设置结果
        """
        array_index = index - 1
        self.rating_status[array_index] = {
            'status': Config.RATING_STATUS_PROCESSING,
            'user_id': user_id,
            'timestamp': pd.Timestamp.now()
        }
        return {"success": True, "status": Config.RATING_STATUS_PROCESSING}

    def release_user_locks(self, user_id: str) -> None:
        """释放指定用户处理的所有评分项
        
        Args:
            user_id: 用户 ID
        """
        for index, status_info in list(self.rating_status.items()):
            if status_info.get('user_id') == user_id and \
               status_info.get('status') == Config.RATING_STATUS_PROCESSING:
                self.rating_status[index] = {
                    'status': Config.RATING_STATUS_IDLE,
                    'user_id': None,
                    'timestamp': pd.Timestamp.now()
                }

    def get_user_statistics(self, user_id: str) -> Dict[str, Any]:
        """获取用户的评分统计信息
        
        Args:
            user_id: 用户 ID
            
        Returns:
            Dict[str, Any]: 统计信息字典
        """
        total_rated = 0
        side_a_scores_list: List[Dict[str, int]] = []
        side_b_scores_list: List[Dict[str, int]] = []

        for item_data in self.data_manager.get_json_data():
            side_a_eval = item_data.get('sideA', {}).get('evaluation', {})
            side_b_eval = item_data.get('sideB', {}).get('evaluation', {})
            
            # 检查是否是该用户评分的
            if side_a_eval.get('rated_by') == user_id or \
               side_b_eval.get('rated_by') == user_id:
                total_rated += 1
                
                # 收集 Side A 评分
                side_a_scores = side_a_eval.get('scores', {})
                if all(k in side_a_scores for k in Config.RATING_DIMENSIONS):
                    side_a_scores_list.append(side_a_scores)
                
                # 收集 Side B 评分
                side_b_scores = side_b_eval.get('scores', {})
                if all(k in side_b_scores for k in Config.RATING_DIMENSIONS):
                    side_b_scores_list.append(side_b_scores)

        return {
            'user_id': user_id,
            'total_rated': total_rated,
            'side_a_avg_scores': self._calc_avg_scores(side_a_scores_list),
            'side_b_avg_scores': self._calc_avg_scores(side_b_scores_list),
            'side_a_distribution': self._calc_score_distribution(side_a_scores_list),
            'side_b_distribution': self._calc_score_distribution(side_b_scores_list)
        }

    def _calc_avg_scores(self, scores_list: List[Dict[str, int]]) -> Dict[str, float]:
        """计算平均分
        
        Args:
            scores_list: 评分列表
            
        Returns:
            Dict[str, float]: 各维度平均分
        """
        if not scores_list:
            return {dim: 0.0 for dim in Config.RATING_DIMENSIONS}
        
        avg = {
            dim: sum(s[dim] for s in scores_list) / len(scores_list)
            for dim in Config.RATING_DIMENSIONS
        }
        return {k: round(v, 2) for k, v in avg.items()}

    def _calc_score_distribution(
        self, 
        scores_list: List[Dict[str, int]]
    ) -> Dict[str, Dict[int, int]]:
        """计算评分分布
        
        Args:
            scores_list: 评分列表
            
        Returns:
            Dict[str, Dict[int, int]]: 各维度的评分分布
        """
        distribution = {
            dim: {score: 0 for score in range(Config.RATING_MIN_SCORE, Config.RATING_MAX_SCORE + 1)}
            for dim in Config.RATING_DIMENSIONS
        }
        
        for scores in scores_list:
            for dim in Config.RATING_DIMENSIONS:
                value = scores.get(dim)
                if value in distribution[dim]:
                    distribution[dim][value] += 1
        
        return distribution


# ==================== 观察者模式 - Excel 保存 ====================

class ExcelSaveObserver(Observer):
    """Excel 保存观察者
    
    监听评分保存事件，将数据导出到 Excel 文件。
    """

    def __init__(self, excel_file: Path = None) -> None:
        """初始化 Excel 观察者
        
        Args:
            excel_file: Excel 文件路径，默认为配置中的路径
        """
        self.excel_file = excel_file or Config.EXCEL_FILE
        self._initialize_excel()

    def _initialize_excel(self) -> None:
        """初始化 Excel 文件"""
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.title = "评分记录"
            self._create_headers(ws)
            wb.save(self.excel_file)

    def _create_headers(self, ws: Any) -> None:
        """创建表头
        
        Args:
            ws: Excel 工作表对象
        """
        headers = [
            "评分对 ID", "病例号", "模型",
            "A 侧 - 医学准确性", "A 侧 - 关键要点", "A 侧 - 逻辑性", "A 侧 - 评论",
            "B 侧 - 医学准确性", "B 侧 - 关键要点", "B 侧 - 逻辑性", "B 侧 - 评论",
            "评分人员"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # 设置列宽
        column_widths = [12, 12, 12, 15, 15, 15, 30, 15, 15, 15, 30, 12]
        for col, width in enumerate(column_widths, 1):
            col_letter = chr(64 + col) if col <= 26 else chr(64 + (col - 26))
            ws.column_dimensions[col_letter].width = width

    def _find_row(self, ws: Any, pair_id: str) -> int:
        """查找指定评分对 ID 的行号
        
        Args:
            ws: Excel 工作表对象
            pair_id: 评分对 ID
            
        Returns:
            int: 行号，未找到返回 -1
        """
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == pair_id:
                return row
        return -1

    def _save_to_excel(self, data: Dict[str, Any]) -> None:
        """保存数据到 Excel
        
        Args:
            data: 评分数据字典
        """
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.title = "评分记录"
            self._create_headers(ws)

        pair_id = data.get('pair_id')
        if not pair_id:
            return

        row = self._find_row(ws, pair_id)

        row_data = [
            pair_id,
            data.get('case_number', ''),
            data.get('model', ''),
            data.get('side_a_accuracy', ''),
            data.get('side_a_recall', ''),
            data.get('side_a_logic', ''),
            data.get('side_a_comment', ''),
            data.get('side_b_accuracy', ''),
            data.get('side_b_recall', ''),
            data.get('side_b_logic', ''),
            data.get('side_b_comment', ''),
            data.get('rated_by_name', '')
        ]

        if row == -1:
            row = ws.max_row + 1
        
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row, column=col, value=value)

        wb.save(self.excel_file)
        logger.info(f"已更新 Excel 记录：{pair_id}")

    def update(self, subject: Subject, event: str, data: Optional[Any] = None) -> None:
        """当主题状态发生变化时调用
        
        Args:
            subject: 主题对象
            event: 事件类型
            data: 事件数据
        """
        if event == 'rating_saved' and data:
            self._save_to_excel(data)


# ==================== 数据处理器 ====================

class DataHandler:
    """数据处理器
    
    统一的数据处理入口，协调各个组件的工作。
    """

    def __init__(self, user_manager: Any) -> None:
        """初始化数据处理器
        
        Args:
            user_manager: 用户管理器实例
        """
        self.data_manager = JsonDataManager()
        self.rating_manager = RatingManager(user_manager, self.data_manager)
        self.excel_observer = ExcelSaveObserver()
        self.data_manager.attach(self.excel_observer)

    def load_json_data(self, app: Optional[Any] = None) -> None:
        """加载 JSON 数据
        
        Args:
            app: Flask 应用实例，用于日志记录
        """
        json_file = Config.get_json_data_file()
        self.data_manager.load_json_data(json_file, app)

    def get_all_ratings(self) -> List[Dict[str, Any]]:
        """获取所有评分项
        
        Returns:
            List[Dict[str, Any]]: 评分项列表
        """
        return self.rating_manager.get_all_ratings()

    def save_rating(self, data: Dict[str, Any], app: Optional[Any] = None) -> Dict[str, Any]:
        """保存评分
        
        Args:
            data: 评分数据
            app: Flask 应用实例
            
        Returns:
            Dict[str, Any]: 保存结果
        """
        return self.rating_manager.save_rating(data, app)

    def get_rating_last_modified(self, index: int) -> Optional[Dict[str, Any]]:
        """获取评分项的最后修改信息
        
        Args:
            index: 评分项索引
            
        Returns:
            Optional[Dict[str, Any]]: 最后修改信息
        """
        return self.rating_manager.get_rating_last_modified(index)

    def get_rating_history(self, index: int) -> List[Dict[str, Any]]:
        """获取评分项的修改历史
        
        Args:
            index: 评分项索引
            
        Returns:
            List[Dict[str, Any]]: 修改历史列表
        """
        return self.rating_manager.get_rating_history(index)

    def set_rating_processing(self, index: int, user_id: str) -> Dict[str, Any]:
        """设置评分项为处理中状态
        
        Args:
            index: 评分项索引
            user_id: 用户 ID
            
        Returns:
            Dict[str, Any]: 设置结果
        """
        return self.rating_manager.set_rating_processing(index, user_id)

    def release_user_locks(self, user_id: str) -> None:
        """释放指定用户处理的所有评分项
        
        Args:
            user_id: 用户 ID
        """
        self.rating_manager.release_user_locks(user_id)

    def get_user_statistics(self, user_id: str) -> Dict[str, Any]:
        """获取用户的评分统计信息
        
        Args:
            user_id: 用户 ID
            
        Returns:
            Dict[str, Any]: 统计信息
        """
        return self.rating_manager.get_user_statistics(user_id)

    def switch_parser_strategy(self, strategy_type: str) -> None:
        """切换解析策略
        
        Args:
            strategy_type: 策略类型
        """
        strategy = JsonParserFactory.create_parser(strategy_type)
        self.data_manager.set_parser_strategy(strategy)
