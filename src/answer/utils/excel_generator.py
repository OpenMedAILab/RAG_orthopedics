"""Excel 生成工具模块"""

from openpyxl import Workbook
from openpyxl.styles import Alignment
from pathlib import Path
from collections import defaultdict


def _get_column_letter(col: int) -> str:
    """获取列字母"""
    if col <= 26:
        return chr(64 + col)
    return f"{chr(64 + col // 26)}{chr(64 + col % 26)}"


def create_model_evaluation_excel(all_results, output_path):
    """创建模型评估 Excel 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "模型评分"

    evaluator_models = list(
        set([r["evaluator_model"] for r in all_results if r["type"] == "evaluation"])
    )

    headers = ["编号", "病例号", "被评估模型", "模式", "PDF 链接"]

    for model in sorted(evaluator_models):
        headers.extend(
            [
                f"{model}-医学准确性 [1-10 分]",
                f"{model}-关键要点召回率 [1-10 分]",
                f"{model}-逻辑完整性 [1-10 分]",
            ]
        )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    output_results = [r for r in all_results if r["type"] == "output"]

    grouped_outputs = defaultdict(list)
    for output in output_results:
        grouped_outputs[
            (output["case_num"], output["model"], output["rag_status"])
        ].append(output)

    eval_results = [r for r in all_results if r["type"] == "evaluation"]
    eval_dict = {}
    for eval_result in eval_results:
        key = (
            eval_result["case_num"],
            eval_result["evaluated_model"],
            eval_result["evaluated_rag_status"],
            eval_result["evaluator_model"],
        )
        eval_dict[key] = eval_result

    row_idx = 2

    for idx, ((case_num, model, rag_status), outputs) in enumerate(
        grouped_outputs.items(), 1
    ):
        output = outputs[0]

        ws.cell(row=row_idx, column=1, value=idx)
        ws.cell(row=row_idx, column=2, value=f"Case {output['case_num']}")
        ws.cell(row=row_idx, column=3, value=output["model"])
        ws.cell(row=row_idx, column=4, value=output["rag_status"])

        pdf_path = output.get("pdf_path", "")
        if pdf_path:
            pdf_link = f'=HYPERLINK("{pdf_path}", "{Path(pdf_path).name}")'
        else:
            pdf_link = "PDF 未生成"
        ws.cell(row=row_idx, column=5, value=pdf_link)

        col_idx = 6
        for evaluator_model in sorted(evaluator_models):
            eval_key = (
                output["case_num"],
                output["model"],
                output["rag_status"],
                evaluator_model,
            )
            if eval_key in eval_dict:
                eval_result = eval_dict[eval_key]
                ws.cell(row=row_idx, column=col_idx, value=eval_result["medical_accuracy"])
                ws.cell(row=row_idx, column=col_idx + 1, value=eval_result["key_point_recall"])
                ws.cell(row=row_idx, column=col_idx + 2, value=eval_result["logical_completeness"])

            col_idx += 3

        row_idx += 1

    for col in range(1, len(headers) + 1):
        col_letter = _get_column_letter(col)
        ws.column_dimensions[col_letter].width = 15

    wb.save(output_path)


def create_empty_evaluation_template(all_results, output_path: str):
    """创建人工评估模板 Excel 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "人工评分"

    headers = [
        "编号",
        "病例号",
        "被评估模型",
        "模式",
        "PDF 链接",
        "医学准确性 [1-10 分]",
        "关键要点召回率 [1-10 分]",
        "逻辑完整性 [1-10 分]",
        "人工评分员",
        "备注",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_idx = 2
    output_results = [r for r in all_results if r["type"] == "output"]
    for idx, result in enumerate(output_results, 1):
        ws.cell(row=row_idx, column=1, value=idx)
        ws.cell(row=row_idx, column=2, value=f"Case {result['case_num']}")
        ws.cell(row=row_idx, column=3, value=result["model"])
        ws.cell(row=row_idx, column=4, value=result["rag_status"])

        pdf_path = result.get("pdf_path", "")
        if pdf_path:
            pdf_link = f'=HYPERLINK("{pdf_path}", "{Path(pdf_path).name}")'
        else:
            pdf_link = "PDF 未生成"
        ws.cell(row=row_idx, column=5, value=pdf_link)

        for col in range(6, 11):
            ws.cell(row=row_idx, column=col, value="")

        row_idx += 1

    for col in range(1, len(headers) + 1):
        col_letter = _get_column_letter(col)
        ws.column_dimensions[col_letter].width = 15

    wb.save(output_path)
